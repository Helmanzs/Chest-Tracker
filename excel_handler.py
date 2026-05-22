"""
excel_handler.py
----------------
Exports DB data to a .xlsx file on demand.

Note: item price loading is handled by prices_config.py, not here.
"""

from __future__ import annotations

from datetime import datetime

import openpyxl
import pandas as pd


def export_to_excel(
    chest_type: str,
    loot_rows: list[dict],
    drop_rates: dict[str, float] | None = None,
    column_order: list[str] | None = None,
    output_path: str | None = None,
) -> str:
    """
    Export loot data fetched from Supabase to a .xlsx file.

    Parameters
    ----------
    chest_type   : used for the sheet name and default filename
    loot_rows    : list of dicts with keys: chest_id, recorded_at, item_name, quantity
    drop_rates   : {item_name: drop_pct} — written to a second sheet if provided
    column_order : item column order to preserve from the viewer (optional)
    output_path  : explicit save path; if None a timestamped filename is generated

    Returns the path the file was saved to.
    """
    if not loot_rows:
        raise ValueError("No data to export")

    # Pivot: one row per chest_id, one column per item
    df = pd.DataFrame(loot_rows)
    pivot = df.pivot_table(
        index=["chest_id", "recorded_at"],
        columns="item_name",
        values="quantity",
        aggfunc="sum",
        fill_value=0,
    ).reset_index()
    pivot.columns.name = None

    # Reorder columns to match viewer display order
    if column_order:
        meta_cols = [c for c in ["chest_id", "recorded_at"] if c in pivot.columns]
        ordered_items = [c for c in column_order if c in pivot.columns and c not in meta_cols]
        remaining = [c for c in pivot.columns if c not in meta_cols and c not in ordered_items]
        pivot = pivot[meta_cols + ordered_items + remaining]

    pivot.insert(0, "#", range(1, len(pivot) + 1))

    # Build output path
    if output_path is None:
        safe_type = chest_type.replace("'", "").replace(" ", "_")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"{safe_type}_export_{ts}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = chest_type[:31]  # type: ignore[union-attr]

    # Write loot sheet header + data
    for col_idx, col_name in enumerate(pivot.columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)  # type: ignore[union-attr]
    for row_idx, row in enumerate(pivot.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)  # type: ignore[union-attr]

    # Second sheet: drop rates
    if drop_rates:
        ws2 = wb.create_sheet(title="Drop Rates")
        ws2.cell(row=1, column=1, value="Item")
        ws2.cell(row=1, column=2, value="Drop Rate %")

        def _rate_sort_key(item: str) -> tuple[float, int]:
            rate = drop_rates.get(item, 0.0)
            order_pos = column_order.index(item) if column_order and item in column_order else 9999
            return (-rate, order_pos)

        meta = {"#", "chest_id", "recorded_at"}
        item_cols = [c for c in pivot.columns if c not in meta]
        for row_idx, item in enumerate(sorted(item_cols, key=_rate_sort_key), start=2):
            rate = drop_rates.get(item)
            ws2.cell(row=row_idx, column=1, value=item)
            if rate is None:
                ws2.cell(row=row_idx, column=2, value="unknown")
            else:
                ws2.cell(row=row_idx, column=2, value=round(rate, 1))

    wb.save(output_path)
    return output_path
